import streamlit as st
import requests
import json
import pandas as pd
import numpy as np
from scipy.stats import poisson, norm
import openpyxl
from datetime import datetime
import os
from dotenv import load_dotenv

# Carregar variáveis de ambiente
load_dotenv()
API_KEY = os.getenv("API_KEY", "f8004fe5cca0e75109a44ae6b4cdd9a2")
API_BASE_URL = "https://api-football-v1.p.rapidapi.com/v3"
HEADERS = {
    "x-rapidapi-key": API_KEY,
    "x-rapidapi-host": "api-football-v1.p.rapidapi.com"
}

# Carregar pesos das competições
def load_weights():
    try:
        with open("pesos.json", "r") as f:
            return json.load(f)
    except FileNotFoundError:
        return {
            "Champions League": 0.95,
            "Europa League": 0.85,
            "Libertadores": 0.80,
            "Sul-Americana": 0.72,
            "Premier League": 0.90,
            "La Liga": 0.87,
            "Brasileirão A": 0.78,
            "Brasileirão B": 0.65,
            "Copa do Brasil": 0.65,
            "Estaduais Fortes": 0.58,
            "Estaduais Fracos": 0.55,
            "MLS": 0.70,
            "Liga MX": 0.72,
            "Saudi Pro League": 0.74
        }

# Função para buscar times
def search_team(team_name, season):
    url = f"{API_BASE_URL}/teams"
    params = {"search": team_name, "season": season}
    response = requests.get(url, headers=HEADERS, params=params)
    if response.status_code == 200:
        return response.json().get("response", [])
    return []

# Função para buscar jogos
def get_team_games(team_id, season, home=True, limit=10):
    url = f"{API_BASE_URL}/fixtures"
    params = {
        "team": team_id,
        "season": season,
        "last": limit,
        "status": "FT"
    }
    if home:
        params["venue"] = "home"
    else:
        params["venue"] = "away"
    response = requests.get(url, headers=HEADERS, params=params)
    if response.status_code == 200:
        return response.json().get("response", [])
    return []

# Função para buscar estatísticas de um jogo
def get_game_stats(fixture_id):
    url = f"{API_BASE_URL}/fixtures/statistics"
    params = {"fixture": fixture_id}
    response = requests.get(url, headers=HEADERS, params=params)
    if response.status_code == 200:
        return response.json().get("response", [])
    return []

# Função para buscar competições de um time
def get_team_leagues(team_id, season):
    url = f"{API_BASE_URL}/leagues"
    params = {"team": team_id, "season": season}
    response = requests.get(url, headers=HEADERS, params=params)
    if response.status_code == 200:
        return response.json().get("response", [])
    return []

# Função para calcular médias
def calculate_averages(games, team_id, weights):
    stats = {
        "goals_scored": [], "goals_conceded": [],
        "shots": [], "shots_on_target": [],
        "corners": [], "possession": [],
        "offsides": [], "fouls_committed": [], "fouls_suffered": [],
        "yellow_cards": [], "red_cards": [],
        "passes_accurate": [], "passes_missed": [],
        "xg": [], "xga": [], "free_kicks": []
    }
    weighted_stats = stats.copy()

    for game in games:
        game_stats = get_game_stats(game["fixture"]["id"])
        if not game_stats:
            continue

        team_stats = next((s for s in game_stats if s["team"]["id"] == team_id), None)
        opponent_stats = next((s for s in game_stats if s["team"]["id"] != team_id), None)
        if not team_stats or not opponent_stats:
            continue

        # Extrair estatísticas
        team_data = {k: 0 for k in stats.keys()}
        for stat in team_stats["statistics"]:
            if stat["type"] == "Goals":
                team_data["goals_scored"] = stat["value"] or 0
            elif stat["type"] == "Shots total":
                team_data["shots"] = stat["value"] or 0
            elif stat["type"] == "Shots on Goal":
                team_data["shots_on_target"] = stat["value"] or 0
            elif stat["type"] == "Corner Kicks":
                team_data["corners"] = stat["value"] or 0
            elif stat["type"] == "Ball Possession":
                team_data["possession"] = float(stat["value"].replace("%", "")) if stat["value"] else 0
            elif stat["type"] == "Offsides":
                team_data["offsides"] = stat["value"] or 0
            elif stat["type"] == "Fouls":
                team_data["fouls_committed"] = stat["value"] or 0
            elif stat["type"] == "Yellow Cards":
                team_data["yellow_cards"] = stat["value"] or 0
            elif stat["type"] == "Red Cards":
                team_data["red_cards"] = stat["value"] or 0
            elif stat["type"] == "Passes accurate":
                team_data["passes_accurate"] = stat["value"] or 0
            elif stat["type"] == "Passes missed":
                team_data["passes_missed"] = stat["value"] or 0
            elif stat["type"] == "expected_goals":
                team_data["xg"] = float(stat["value"]) if stat["value"] else 0
            elif stat["type"] == "expected_goals_against":
                team_data["xga"] = float(stat["value"]) if stat["value"] else 0
            elif stat["type"] == "Free Kicks":
                team_data["free_kicks"] = stat["value"] or 0

        opponent_id = opponent_stats["team"]["id"]
        leagues = get_team_leagues(opponent_id, game["league"]["season"])
        max_weight = 0.50
        for league in leagues:
            league_name = league["league"]["name"]
            weight = weights.get(league_name, 0.50)
            max_weight = max(max_weight, weight)

        # Média simples
        for key in stats:
            stats[key].append(team_data[key])

        # Média ponderada
        for key in weighted_stats:
            if "conceded" in key or "suffered" in key:
                weighted_stats[key].append(team_data[key] / max_weight)
            else:
                weighted_stats[key].append(team_data[key] * max_weight)

    simple_averages = {k: np.mean(v) if v else 0 for k, v in stats.items()}
    weighted_averages = {k: np.mean(v) if v else 0 for k, v in weighted_stats.items()}
    return simple_averages, weighted_averages

# Função para prever estatísticas
def predict_stats(team_a_simple, team_a_weighted, team_b_simple, team_b_weighted):
    predicted_stats = {}
    confidence_intervals = {}
    for stat in team_a_weighted.keys():
        # Previsão para Time A
        a_pred = (team_a_weighted[stat] + team_b_weighted[stat.replace("scored", "conceded").replace("committed", "suffered")]) * 1.2 / 2
        # Previsão para Time B
        b_pred = (team_b_weighted[stat] + team_a_weighted[stat.replace("scored", "conceded").replace("committed", "suffered")]) / 1.2 / 2
        predicted_stats[stat] = {"team_a": a_pred, "team_b": b_pred}

        # Intervalo de confiança (85%)
        a_std = np.std([team_a_weighted[stat], team_b_weighted[stat.replace("scored", "conceded").replace("committed", "suffered")]])
        b_std = np.std([team_b_weighted[stat], team_a_weighted[stat.replace("scored", "conceded").replace("committed", "suffered")]])
        z = norm.ppf(0.925)  # 85% CI
        confidence_intervals[stat] = {
            "team_a": (a_pred - z * a_std / np.sqrt(2), a_pred + z * a_std / np.sqrt(2)),
            "team_b": (b_pred - z * b_std / np.sqrt(2), b_pred + z * b_std / np.sqrt(2))
        }
    return predicted_stats, confidence_intervals

# Função para prever placar
def predict_score(team_a_weighted, team_b_weighted):
    lambda_a = (team_a_weighted["goals_scored"] + team_b_weighted["goals_conceded"]) * 1.2 / 2
    lambda_b = (team_b_weighted["goals_scored"] + team_a_weighted["goals_conceded"]) / 1.2 / 2

    max_goals = 10
    prob_matrix = np.zeros((max_goals, max_goals))
    for i in range(max_goals):
        for j in range(max_goals):
            prob_matrix[i, j] = poisson.pmf(i, lambda_a) * poisson.pmf(j, lambda_b)

    most_likely_score = np.unravel_index(np.argmax(prob_matrix), prob_matrix.shape)
    win_prob = np.sum(prob_matrix[np.where(np.arange(max_goals)[:, None] > np.arange(max_goals)[None, :])])
    draw_prob = np.sum(np.diag(prob_matrix))
    loss_prob = np.sum(prob_matrix[np.where(np.arange(max_goals)[:, None] < np.arange(max_goals)[None, :])])

    ci_a = (poisson.ppf(0.075, lambda_a), poisson.ppf(0.925, lambda_a))
    ci_b = (poisson.ppf(0.075, lambda_b), poisson.ppf(0.925, lambda_b))

    return {
        "score": f"{most_likely_score[0]} x {most_likely_score[1]}",
        "probs": {"win": win_prob, "draw": draw_prob, "loss": loss_prob},
        "ci": {"team_a": ci_a, "team_b": ci_b}
    }

# Função para buscar odds
def get_odds(fixture_id):
    url = f"{API_BASE_URL}/odds"
    params = {"fixture": fixture_id}
    response = requests.get(url, headers=HEADERS, params=params)
    if response.status_code == 200:
        return response.json().get("response", [])
    return []

# Função para comparar odds e previsões
def compare_odds(predicted_stats, score_pred, odds):
    value_bets = []
    for market in odds:
        market_name = market["bookmaker"]["name"]
        for bet in market["bets"]:
            bet_name = bet["name"]
            for value in bet["values"]:
                odd = float(value["odd"])
                implied_prob = 1 / odd
                predicted_prob = 0
                if bet_name == "Match Winner":
                    if value["value"] == "Home":
                        predicted_prob = score_pred["probs"]["win"]
                    elif value["value"] == "Draw":
                        predicted_prob = score_pred["probs"]["draw"]
                    elif value["value"] == "Away":
                        predicted_prob = score_pred["probs"]["loss"]
                elif bet_name == "Both Teams To Score":
                    if value["value"] == "Yes":
                        predicted_prob = poisson.pmf(1, predicted_stats["goals_scored"]["team_a"]) * poisson.pmf(1, predicted_stats["goals_scored"]["team_b"])
                if predicted_prob > implied_prob:
                    value_bets.append({
                        "market": market_name,
                        "bet": bet_name,
                        "value": value["value"],
                        "odd": odd,
                        "implied_prob": implied_prob,
                        "predicted_prob": predicted_prob
                    })
    return value_bets

# Função para exportar resultados
def export_results(team_a, team_b, simple_a, weighted_a, simple_b, weighted_b, predicted_stats, confidence_intervals, score_pred, value_bets):
    wb = openpyxl.Workbook()
    
    # Aba de médias
    ws = wb.active
    ws.title = "Médias"
    ws.append(["Time", "Estatística", "Média Simples", "Média Ponderada"])
    for stat in simple_a.keys():
        ws.append([team_a["team"]["name"], stat, simple_a[stat], weighted_a[stat]])
        ws.append([team_b["team"]["name"], stat, simple_b[stat], weighted_b[stat]])

    # Aba de previsões
    ws = wb.create_sheet("Previsões")
    ws.append(["Estatística", f"{team_a['team']['name']} (Prev)", f"{team_a['team']['name']} (IC 85%)", f"{team_b['team']['name']} (Prev)", f"{team_b['team']['name']} (IC 85%)"])
    for stat in predicted_stats.keys():
        ws.append([
            stat,
            predicted_stats[stat]["team_a"],
            f"[{confidence_intervals[stat]['team_a'][0]:.2f}, {confidence_intervals[stat]['team_a'][1]:.2f}]",
            predicted_stats[stat]["team_b"],
            f"[{confidence_intervals[stat]['team_b'][0]:.2f}, {confidence_intervals[stat]['team_b'][1]:.2f}]"
        ])

    # Aba de placar
    ws = wb.create_sheet("Placar Provável")
    ws.append(["Placar", score_pred["score"]])
    ws.append(["Prob. Vitória", score_pred["probs"]["win"]])
    ws.append(["Prob. Empate", score_pred["probs"]["draw"]])
    ws.append(["Prob. Derrota", score_pred["probs"]["loss"]])
    ws.append([f"IC Gols {team_a['team']['name']}", f"[{score_pred['ci']['team_a'][0]:.2f}, {score_pred['ci']['team_a'][1]:.2f}]"])
    ws.append([f"IC Gols {team_b['team']['name']}", f"[{score_pred['ci']['team_b'][0]:.2f}, {score_pred['ci']['team_b'][1]:.2f}]"])

    # Aba de odds
    ws = wb.create_sheet("Odds e Valor")
    ws.append(["Mercado", "Aposta", "Valor", "Odd", "Prob. Implícita", "Prob. Prevista"])
    for bet in value_bets:
        ws.append([bet["market"], bet["bet"], bet["value"], bet["odd"], bet["implied_prob"], bet["predicted_prob"]])

    filename = f"previsao_{team_a['team']['name']}_vs_{team_b['team']['name']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)
    return filename

# Função principal
def main():
    st.set_page_config(page_title="Previsão de Partidas de Futebol", layout="wide")
    st.title("Previsão Estatística de Partidas de Futebol")

    weights = load_weights()
    tabs = st.tabs([
        "Seleção de Times",
        "Jogos Analisados",
        "Médias",
        "Estatísticas Previstas",
        "Placar Provável",
        "Odds x Previsão",
        "Exportar"
    ])

    # Aba 1: Seleção de Times
    with tabs[0]:
        col1, col2 = st.columns(2)
        with col1:
            team_a_name = st.text_input("Time A (Mandante)", placeholder="Digite o nome do time")
            season_a = st.selectbox("Temporada Time A", list(range(2020, 2026)), index=3)
        with col2:
            team_b_name = st.text_input("Time B (Visitante)", placeholder="Digite o nome do time")
            season_b = st.selectbox("Temporada Time B", list(range(2020, 2026)), index=3)

        if st.button("Buscar Times"):
            teams_a = search_team(team_a_name, season_a)
            teams_b = search_team(team_b_name, season_b)
            st.session_state["teams_a"] = teams_a
            st.session_state["teams_b"] = teams_b

        if "teams_a" in st.session_state and "teams_b" in st.session_state:
            st.subheader("Selecione os Times")
            team_a_selected = st.selectbox("Time A", [f"{t['team']['name']} ({t['team']['country']})" for t in st.session_state["teams_a"]])
            team_b_selected = st.selectbox("Time B", [f"{t['team']['name']} ({t['team']['country']})" for t in st.session_state["teams_b"]])
            if st.button("Confirmar Seleção"):
                st.session_state["team_a"] = next(t for t in st.session_state["teams_a"] if f"{t['team']['name']} ({t['team']['country']})" == team_a_selected)
                st.session_state["team_b"] = next(t for t in st.session_state["teams_b"] if f"{t['team']['name']} ({t['team']['country']})" == team_b_selected)
                st.success("Times selecionados com sucesso!")

    # Aba 2: Jogos Analisados
    with tabs[1]:
        if "team_a" in st.session_state and "team_b" in st.session_state:
            team_a_id = st.session_state["team_a"]["team"]["id"]
            team_b_id = st.session_state["team_b"]["team"]["id"]
            games_a = get_team_games(team_a_id, season_a, home=True)
            games_b = get_team_games(team_b_id, season_b, home=False)
            st.write("Jogos do Time A (Mandante):")
            for game in games_a:
                with st.expander(f"{game['teams']['home']['name']} vs {game['teams']['away']['name']}"):
                    st.json(get_game_stats(game["fixture"]["id"]))
            st.write("Jogos do Time B (Visitante):")
            for game in games_b:
                with st.expander(f"{game['teams']['home']['name']} vs {game['teams']['away']['name']}"):
                    st.json(get_game_stats(game["fixture"]["id"]))

    # Aba 3: Médias
    with tabs[2]:
        if "team_a" in st.session_state and "team_b" in st.session_state:
            team_a_id = st.session_state["team_a"]["team"]["id"]
            team_b_id = st.session_state["team_b"]["team"]["id"]
            games_a = get_team_games(team_a_id, season_a, home=True)
            games_b = get_team_games(team_b_id, season_b, home=False)
            simple_a, weighted_a = calculate_averages(games_a, team_a_id, weights)
            simple_b, weighted_b = calculate_averages(games_b, team_b_id, weights)
            st.session_state["simple_a"] = simple_a
            st.session_state["weighted_a"] = weighted_a
            st.session_state["simple_b"] = simple_b
            st.session_state["weighted_b"] = weighted_b

            df_a = pd.DataFrame({
                "Estatística": simple_a.keys(),
                "Média Simples": simple_a.values(),
                "Média Ponderada": weighted_a.values()
            })
            df_b = pd.DataFrame({
                "Estatística": simple_b.keys(),
                "Média Simples": simple_b.values(),
                "Média Ponderada": weighted_b.values()
            })
            st.write(f"Médias do Time A ({st.session_state['team_a']['team']['name']}):")
            st.dataframe(df_a)
            st.write(f"Médias do Time B ({st.session_state['team_b']['team']['name']}):")
            st.dataframe(df_b)

    # Aba 4: Estatísticas Previstas
    with tabs[3]:
        if "simple_a" in st.session_state:
            predicted_stats, confidence_intervals = predict_stats(
                st.session_state["simple_a"], st.session_state["weighted_a"],
                st.session_state["simple_b"], st.session_state["weighted_b"]
            )
            st.session_state["predicted_stats"] = predicted_stats
            st.session_state["confidence_intervals"] = confidence_intervals
            df_pred = pd.DataFrame([
                {
                    "Estatística": stat,
                    f"{st.session_state['team_a']['team']['name']}": pred["team_a"],
                    f"{st.session_state['team_a']['team']['name']} (IC 85%)": f"[{confidence_intervals[stat]['team_a'][0]:.2f}, {confidence_intervals[stat]['team_a'][1]:.2f}]",
                    f"{st.session_state['team_b']['team']['name']}": pred["team_b"],
                    f"{st.session_state['team_b']['team']['name']} (IC 85%)": f"[{confidence_intervals[stat]['team_b'][0]:.2f}, {confidence_intervals[stat]['team_b'][1]:.2f}]"
                }
                for stat, pred in predicted_stats.items()
            ])
            st.dataframe(df_pred)

    # Aba 5: Placar Provável
    with tabs[4]:
        if "weighted_a" in st.session_state:
            score_pred = predict_score(st.session_state["weighted_a"], st.session_state["weighted_b"])
            st.session_state["score_pred"] = score_pred
            st.write(f"Placar mais provável: {score_pred['score']}")
            st.write(f"Probabilidade de Vitória: {score_pred['probs']['win']:.2%}")
            st.write(f"Probabilidade de Empate: {score_pred['probs']['draw']:.2%}")
            st.write(f"Probabilidade de Derrota: {score_pred['probs']['loss']:.2%}")
            st.write(f"Intervalo de Confiança (Gols {st.session_state['team_a']['team']['name']}): [{score_pred['ci']['team_a'][0]:.2f}, {score_pred['ci']['team_a'][1]:.2f}]")
            st.write(f"Intervalo de Confiança (Gols {st.session_state['team_b']['team']['name']}): [{score_pred['ci']['team_b'][0]:.2f}, {score_pred['ci']['team_b'][1]:.2f}]")

    # Aba 6: Odds x Previsão
    with tabs[5]:
        if "predicted_stats" in st.session_state:
            # Simulação de fixture_id (em produção, obter via API)
            odds = get_odds(12345)  # Substituir por chamada real
            value_bets = compare_odds(st.session_state["predicted_stats"], st.session_state["score_pred"], odds)
            st.session_state["value_bets"] = value_bets
            if value_bets:
                df_odds = pd.DataFrame(value_bets)
                st.dataframe(df_odds)
            else:
                st.write("Nenhuma aposta de valor encontrada.")

    # Aba 7: Exportar
    with tabs[6]:
        if "team_a" in st.session_state and "predicted_stats" in st.session_state:
            if st.button("Exportar Resultados"):
                filename = export_results(
                    st.session_state["team_a"], st.session_state["team_b"],
                    st.session_state["simple_a"], st.session_state["weighted_a"],
                    st.session_state["simple_b"], st.session_state["weighted_b"],
                    st.session_state["predicted_stats"], st.session_state["confidence_intervals"],
                    st.session_state["score_pred"], st.session_state["value_bets"]
                )
                with open(filename, "rb") as f:
                    st.download_button("Baixar .xlsx", f, file_name=filename)
                st.success("Resultados exportados com sucesso!")

if __name__ == "__main__":
    main()